import os 
import xlrd
import openpyxl


def open_xls_as_xlsx(xls_path, xlsx_path):
    '''
    将excel中的xls样式改为xlsx
    '''
    book = xlrd.open_workbook(xls_path)
    index = 0
    nrows = 0
    ncols = 0
    sheet = book.sheet_by_index(0)
    while nrows*ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1
        
    book_new = openpyxl.Workbook()
    book_new.create_sheet('sheet1',0)
    sheet_new = book_new['sheet1']

    
    # print(help(sheet_new))
    # print(dir(sheet_new))
    
    for row in range(0,nrows):
        for col in range(0,ncols):
            sheet_new.cell(row=row+1,column=col+1).value = sheet.cell_value(row,col)
    book_new.save(xlsx_path)

def processexcel(f,quiz):
    workbook = openpyxl.load_workbook(f)
    sheet = workbook.active
    # print(sheet)
    for i in sheet.iter_rows(min_row=1,max_row = 9,min_col=1,max_col=9):
        s = []
        for j in i:
            v = j.value
            if v == '' or v is None:
                v = '0'
            s.append(int(v))
        quiz.append(s)
 

def getquizfromexcel(quiz):
    current_file_path = os.path.dirname(__file__)
    # print(current_file_path)
    # # path = r"./"#必须cd到本文件同一文件夹内
    
    # print(f_path)
    
    for f1 in os.listdir(current_file_path):
        # print(f1)
        if f1.endswith('.xls'):
            # print(f_path)
            f1 = os.path.join(current_file_path,f1)
            f_path = os.path.join(current_file_path,'source.xlsx')            
            open_xls_as_xlsx(f1,f_path)
            processexcel(f_path,quiz)
            break
            

    # for f in os.listdir(path):
    #     if f.endswith('.xlsx') :
    #         # print(f)
    #         # 处理表格，输出quiz数组
    #         processexcel(f,quiz)
    #         break
    else:
        print('No such file.')
    
        
if __name__ == '__main__':
    quiz = []
    getquizfromexcel(quiz)
    for item in quiz:
        print(item)